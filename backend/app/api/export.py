import io
import csv
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_
from app.database import get_db
from app.models.personnel import HonorariosRecord, ContrataRecord, PlantaRecord
from typing import Optional

router = APIRouter()


def _build_filters(M, municipality_code, area, year, months, convenios, search_text):
    clauses = [M.municipality_code == municipality_code, M.area == area, M.year == year]
    if months:
        clauses.append(M.month.in_([int(m) for m in months.split(",")]))
    if convenios and hasattr(M, "convenio"):
        clauses.append(M.convenio.in_(convenios.split(",")))
    if search_text:
        p = f"%{search_text}%"
        name_col = M.nombre
        search_cols = [name_col.ilike(p)]
        if hasattr(M, "cargo"):
            search_cols.append(M.cargo.ilike(p))
        if hasattr(M, "descripcion_funcion"):
            search_cols.append(M.descripcion_funcion.ilike(p))
        if hasattr(M, "calificacion_profesional"):
            search_cols.append(M.calificacion_profesional.ilike(p))
        clauses.append(or_(*search_cols))
    return and_(*clauses)


@router.get("/csv")
async def export_csv(
    municipality_code: str = Query(...),
    area: str = Query("Salud"),
    year: int = Query(2025),
    contract_type: str = Query("HONORARIOS"),
    months: Optional[str] = Query(None),
    convenios: Optional[str] = Query(None),
    search_text: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    model_map = {
        "HONORARIOS": HonorariosRecord,
        "CONTRATA": ContrataRecord,
        "PLANTA": PlantaRecord,
    }
    M = model_map.get(contract_type, HonorariosRecord)
    where = _build_filters(M, municipality_code, area, year, months, convenios, search_text)

    q = select(M).where(where).order_by(M.month if hasattr(M, "month") else M.id)
    rows = (await db.execute(q)).scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)

    if contract_type == "HONORARIOS":
        headers = [
            "Mes", "Nombre", "RUT", "Función", "Calificación", "Fecha Inicio",
            "Fecha Término", "Rem. Bruta", "Rem. Líquida", "Monto Total",
            "Observaciones", "Convenio"
        ]
        writer.writerow(headers)
        for r in rows:
            writer.writerow([
                r.month, r.nombre, r.rut, r.descripcion_funcion,
                r.calificacion_profesional, r.fecha_inicio, r.fecha_termino,
                r.remuneracion_bruta, r.remuneracion_liquida, r.monto_total,
                r.observaciones, r.convenio,
            ])
    else:
        headers = [
            "Mes", "Nombre", "RUT", "Grado", "Cargo", "Calificación",
            "Asignaciones", "Rem. Bruta", "Rem. Líquida",
            "Fecha Inicio", "Fecha Término", "Observaciones", "Horas"
        ]
        writer.writerow(headers)
        for r in rows:
            writer.writerow([
                r.month, r.nombre, r.rut, r.grado_eus, r.cargo,
                r.calificacion_profesional, r.asignaciones,
                r.remuneracion_bruta, r.remuneracion_liquida,
                r.fecha_inicio, r.fecha_termino, r.observaciones, r.horas,
            ])

    output.seek(0)
    filename = f"munidata_{municipality_code}_{contract_type}_{year}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/excel")
async def export_excel(
    municipality_code: str = Query(...),
    area: str = Query("Salud"),
    year: int = Query(2025),
    months: Optional[str] = Query(None),
    convenios: Optional[str] = Query(None),
    search_text: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    import openpyxl

    wb = openpyxl.Workbook()

    # Honorarios sheet
    ws_hon = wb.active
    ws_hon.title = "Honorarios"
    hon_headers = [
        "Mes", "Nombre", "RUT", "Función", "Calificación", "Fecha Inicio",
        "Fecha Término", "Rem. Bruta", "Rem. Líquida", "Monto Total",
        "Observaciones", "Convenio"
    ]
    ws_hon.append(hon_headers)

    where_hon = _build_filters(HonorariosRecord, municipality_code, area, year, months, convenios, search_text)
    rows_hon = (await db.execute(select(HonorariosRecord).where(where_hon).order_by(HonorariosRecord.month))).scalars().all()
    for r in rows_hon:
        ws_hon.append([
            r.month, r.nombre, r.rut, r.descripcion_funcion,
            r.calificacion_profesional, r.fecha_inicio, r.fecha_termino,
            r.remuneracion_bruta, r.remuneracion_liquida, r.monto_total,
            r.observaciones, r.convenio,
        ])

    # Contrata sheet
    ws_cont = wb.create_sheet("Contrata")
    cont_headers = [
        "Mes", "Nombre", "RUT", "Grado", "Cargo", "Calificación",
        "Asignaciones", "Rem. Bruta", "Rem. Líquida",
        "Fecha Inicio", "Fecha Término", "Observaciones", "Horas"
    ]
    ws_cont.append(cont_headers)

    where_cont = _build_filters(ContrataRecord, municipality_code, area, year, months, None, search_text)
    rows_cont = (await db.execute(select(ContrataRecord).where(where_cont).order_by(ContrataRecord.month))).scalars().all()
    for r in rows_cont:
        ws_cont.append([
            r.month, r.nombre, r.rut, r.grado_eus, r.cargo,
            r.calificacion_profesional, r.asignaciones,
            r.remuneracion_bruta, r.remuneracion_liquida,
            r.fecha_inicio, r.fecha_termino, r.observaciones, r.horas,
        ])

    # Planta sheet
    ws_pla = wb.create_sheet("Planta")
    ws_pla.append(cont_headers)

    where_pla = _build_filters(PlantaRecord, municipality_code, area, year, months, None, search_text)
    rows_pla = (await db.execute(select(PlantaRecord).where(where_pla).order_by(PlantaRecord.month))).scalars().all()
    for r in rows_pla:
        ws_pla.append([
            r.month, r.nombre, r.rut, r.grado_eus, r.cargo,
            r.calificacion_profesional, r.asignaciones,
            r.remuneracion_bruta, r.remuneracion_liquida,
            r.fecha_inicio, r.fecha_termino, r.observaciones, r.horas,
        ])

    # Control sheet (summary)
    ws_ctrl = wb.create_sheet("Resumen")
    ws_ctrl.append(["Municipio", municipality_code])
    ws_ctrl.append(["Área", area])
    ws_ctrl.append(["Año", year])
    ws_ctrl.append(["Total Honorarios", len(rows_hon)])
    ws_ctrl.append(["Total Contrata", len(rows_cont)])
    ws_ctrl.append(["Total Planta", len(rows_pla)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"munidata_{municipality_code}_{year}_consolidado.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
